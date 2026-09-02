import json
import os
import random
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import winsound
except ImportError:
    winsound = None


APP_NAME = "学生随机点名系统"
APP_VERSION = "3.0.0"
TEMPLATE_NAME = "学生导入模板.xlsx"
AUTO_HIDE_SECONDS = 8


def get_data_directory():
    custom_directory = os.getenv("STUDENT_ROLL_CALL_DATA_DIR")
    local_app_data = os.getenv("LOCALAPPDATA")
    if custom_directory:
        directory = Path(custom_directory)
    elif local_app_data:
        directory = Path(local_app_data) / APP_NAME
    else:
        directory = Path.home() / APP_NAME
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def resource_path(file_name):
    base_path = Path(sys._MEIPASS) if hasattr(sys, "_MEIPASS") else Path(__file__).parent
    return base_path / file_name


def masked_name(name):
    clean_name = name.strip()
    return f"{clean_name[0]}**" if clean_name else "***"


def excel_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def participation_enabled(value):
    text = excel_cell_text(value).lower().replace(" ", "")
    return text not in {"否", "不参与", "no", "false", "0", "停用", "禁用"}


def find_excel_columns(worksheet):
    header_names = {
        "student_no": {"学号", "学生学号", "studentid", "studentno", "studentnumber"},
        "name": {"姓名", "学生姓名", "name", "studentname"},
        "class_name": {"班级", "班别", "class", "classname"},
        "active": {"参与点名", "是否参与", "参与", "active", "enabled"},
    }

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=20, max_col=30), start=1
    ):
        columns = {}
        for column_number, cell in enumerate(row, start=1):
            normalized = excel_cell_text(cell.value).lower().replace(" ", "").replace("_", "")
            for field_name, accepted_names in header_names.items():
                if normalized in accepted_names:
                    columns[field_name] = column_number
        if "student_no" in columns and "name" in columns:
            return row_number, columns
    return None


class StudentDatabase:
    def __init__(self):
        self.data_directory = get_data_directory()
        self.database_path = self.data_directory / "students.db"
        self.backup_directory = self.data_directory / "backups"
        self.backup_directory.mkdir(exist_ok=True)
        self.initialize()

    def connect(self):
        connection = sqlite3.connect(self.database_path)
        connection.row_factory = sqlite3.Row
        return connection

    def initialize(self):
        with self.connect() as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS students (
                    student_no TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            columns = {
                row["name"] for row in connection.execute("PRAGMA table_info(students)")
            }
            if "class_name" not in columns:
                connection.execute(
                    "ALTER TABLE students ADD COLUMN class_name TEXT NOT NULL DEFAULT '默认班级'"
                )
            if "active" not in columns:
                connection.execute(
                    "ALTER TABLE students ADD COLUMN active INTEGER NOT NULL DEFAULT 1"
                )

            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS app_meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS roll_state (
                    class_name TEXT PRIMARY KEY,
                    remaining_json TEXT NOT NULL,
                    round_no INTEGER NOT NULL
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS roll_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id TEXT NOT NULL,
                    class_name TEXT NOT NULL,
                    round_no INTEGER NOT NULL,
                    student_no TEXT NOT NULL,
                    name TEXT NOT NULL,
                    selected_at TEXT NOT NULL
                )
                """
            )

            seeded = connection.execute(
                "SELECT value FROM app_meta WHERE key = 'sample_data_created'"
            ).fetchone()
            if seeded is None:
                connection.executemany(
                    """
                    INSERT OR IGNORE INTO students
                        (student_no, name, class_name, active)
                    VALUES (?, ?, ?, ?)
                    """,
                    [
                        ("S001", "张三", "示例班级", 1),
                        ("S002", "李四", "示例班级", 1),
                        ("S003", "王五", "示例班级", 1),
                        ("S004", "赵六", "示例班级", 1),
                        ("S005", "钱七", "示例班级", 1),
                    ],
                )
                connection.execute(
                    "INSERT INTO app_meta (key, value) VALUES (?, ?)",
                    ("sample_data_created", "1"),
                )

    def get_classes(self):
        with self.connect() as connection:
            rows = connection.execute(
                """
                SELECT class_name, COUNT(*) AS student_count
                FROM students
                WHERE active = 1
                GROUP BY class_name
                ORDER BY class_name
                """
            ).fetchall()
        return [(row["class_name"], row["student_count"]) for row in rows]

    def get_students(self, class_name):
        with self.connect() as connection:
            rows = connection.execute(
                """
                SELECT student_no, name
                FROM students
                WHERE class_name = ? AND active = 1
                ORDER BY student_no
                """,
                (class_name,),
            ).fetchall()
        return [(row["student_no"], row["name"]) for row in rows]

    def get_roll_state(self, class_name):
        with self.connect() as connection:
            row = connection.execute(
                "SELECT remaining_json, round_no FROM roll_state WHERE class_name = ?",
                (class_name,),
            ).fetchone()
        if row is None:
            return None
        try:
            remaining = json.loads(row["remaining_json"])
        except json.JSONDecodeError:
            remaining = []
        return remaining, row["round_no"]

    def save_roll_result(
        self, session_id, class_name, round_no, remaining, student_no, name
    ):
        selected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with self.connect() as connection:
            connection.execute(
                """
                INSERT INTO roll_state (class_name, remaining_json, round_no)
                VALUES (?, ?, ?)
                ON CONFLICT(class_name) DO UPDATE SET
                    remaining_json = excluded.remaining_json,
                    round_no = excluded.round_no
                """,
                (class_name, json.dumps(remaining, ensure_ascii=False), round_no),
            )
            connection.execute(
                """
                INSERT INTO roll_history
                    (session_id, class_name, round_no, student_no, name, selected_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (session_id, class_name, round_no, student_no, name, selected_at),
            )

    def get_last_selected_no(self, class_name):
        with self.connect() as connection:
            row = connection.execute(
                """
                SELECT student_no FROM roll_history
                WHERE class_name = ? ORDER BY id DESC LIMIT 1
                """,
                (class_name,),
            ).fetchone()
        return row["student_no"] if row else None

    def get_session_history(self, session_id):
        with self.connect() as connection:
            rows = connection.execute(
                """
                SELECT selected_at, class_name, round_no, student_no, name
                FROM roll_history
                WHERE session_id = ? ORDER BY id
                """,
                (session_id,),
            ).fetchall()
        return [tuple(row) for row in rows]

    def create_backup(self, prefix="before_import"):
        if not self.database_path.exists():
            return None
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = self.backup_directory / f"{prefix}_{timestamp}.db"
        shutil.copy2(self.database_path, backup_path)
        backups = sorted(self.backup_directory.glob("*.db"), reverse=True)
        for old_backup in backups[10:]:
            old_backup.unlink(missing_ok=True)
        return backup_path

    def replace_students(self, students):
        self.create_backup()
        with self.connect() as connection:
            connection.execute("DELETE FROM students")
            connection.executemany(
                """
                INSERT INTO students (student_no, name, class_name, active)
                VALUES (?, ?, ?, ?)
                """,
                students,
            )
            connection.execute("DELETE FROM roll_state")

    def latest_backup(self):
        backups = sorted(self.backup_directory.glob("*.db"), reverse=True)
        return backups[0] if backups else None

    def restore_latest_backup(self):
        backup = self.latest_backup()
        if backup is None:
            return False
        current_copy = self.backup_directory / (
            "before_restore_" + datetime.now().strftime("%Y%m%d_%H%M%S_%f") + ".db"
        )
        shutil.copy2(self.database_path, current_copy)
        shutil.copy2(backup, self.database_path)
        self.initialize()
        return True


class StudentRollCallApp:
    def __init__(self, root):
        self.root = root
        self.database = StudentDatabase()
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.is_rolling = False
        self.is_fullscreen = False
        self.sound_enabled = tk.BooleanVar(value=True)
        self.selected_class = tk.StringVar()
        self.count_text = tk.StringVar(value="请先导入 Excel 名单")
        self.status_text = tk.StringVar(value="学生名单只能通过 Excel 修改")
        self.auto_hide_job = None
        self.pending_student = None
        self.pending_remaining = []
        self.pending_round_no = 1
        self.roll_step = 0
        self.total_roll_steps = 0
        self.confetti_items = []

        self.configure_window()
        self.configure_styles()
        self.create_widgets()
        self.bind_shortcuts()
        self.refresh_classes()

    def configure_window(self):
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        width = min(900, max(720, screen_width - 40))
        height = min(600, max(520, screen_height - 80))
        x = max(0, (screen_width - width) // 2)
        y = max(0, (screen_height - height) // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.minsize(min(720, width), min(520, height))
        self.root.configure(bg="#0C1728")
        try:
            self.root.iconbitmap(default=str(resource_path("app_icon.ico")))
        except tk.TclError:
            pass

    def configure_styles(self):
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
        style.configure("App.TFrame", background="#0C1728")
        style.configure("Title.TLabel", background="#0C1728", foreground="#FFFFFF",
                        font=("Microsoft YaHei UI", 19, "bold"))
        style.configure("Hint.TLabel", background="#0C1728", foreground="#9CB0C9",
                        font=("Microsoft YaHei UI", 9))
        style.configure("Important.TLabel", background="#0C1728", foreground="#FFD166",
                        font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("Stats.TLabel", background="#0C1728", foreground="#D9E5F3",
                        font=("Microsoft YaHei UI", 10))
        style.configure("Primary.TButton", font=("Microsoft YaHei UI", 15, "bold"),
                        padding=(38, 13))
        style.configure("Tool.TButton", font=("Microsoft YaHei UI", 9), padding=(10, 6))
        style.configure("Status.TLabel", background="#14243B", foreground="#D5E0EE",
                        font=("Microsoft YaHei UI", 9))

    def create_widgets(self):
        main = ttk.Frame(self.root, style="App.TFrame", padding=(24, 20, 24, 12))
        main.pack(fill="both", expand=True)

        header = ttk.Frame(main, style="App.TFrame")
        header.pack(fill="x")
        ttk.Label(header, text=APP_NAME, style="Title.TLabel").pack(side="left")

        self.fullscreen_button = ttk.Button(
            header, text="全屏 F11", command=self.toggle_fullscreen, style="Tool.TButton"
        )
        self.fullscreen_button.pack(side="right")
        self.class_combo = ttk.Combobox(
            header, textvariable=self.selected_class, state="readonly", width=18
        )
        self.class_combo.pack(side="right", padx=(8, 10), pady=4)
        ttk.Label(header, text="班级", style="Hint.TLabel").pack(side="right")
        self.class_combo.bind("<<ComboboxSelected>>", self.on_class_changed)

        notice_row = ttk.Frame(main, style="App.TFrame")
        notice_row.pack(fill="x", pady=(5, 14))
        ttk.Label(
            notice_row,
            text="学生名单只能通过 Excel 修改，操作界面不会显示名单",
            style="Important.TLabel",
        ).pack(side="left")
        ttk.Label(notice_row, textvariable=self.count_text, style="Stats.TLabel").pack(
            side="right"
        )

        stage_frame = tk.Frame(main, bg="#FFFFFF", bd=0, highlightthickness=0)
        stage_frame.pack(fill="both", expand=True)
        self.stage = tk.Canvas(stage_frame, bg="#FFFFFF", highlightthickness=0)
        self.stage.pack(fill="both", expand=True)
        self.stage.bind("<Configure>", self.position_stage_items)

        self.stage_title = self.stage.create_text(
            0, 0, text="准备点名", fill="#607086",
            font=("Microsoft YaHei UI", 15, "bold")
        )
        self.stage_result = self.stage.create_text(
            0, 0, text="按空格键开始", fill="#0E7A6D",
            font=("Microsoft YaHei UI", 44, "bold")
        )
        self.stage_detail = self.stage.create_text(
            0, 0, text="一轮内每名学生只会出现一次", fill="#65758A",
            font=("Microsoft YaHei UI", 11)
        )

        stage_controls = tk.Frame(stage_frame, bg="#FFFFFF")
        stage_controls.pack(side="bottom", fill="x", pady=(0, 18))
        self.roll_button = ttk.Button(
            stage_controls, text="开始点名", command=self.start_roll_call,
            style="Primary.TButton"
        )
        self.roll_button.pack()

        tools = ttk.Frame(main, style="App.TFrame")
        tools.pack(fill="x", pady=(14, 0))
        self.import_button = ttk.Button(
            tools, text="导入 / 更新 Excel", command=self.import_excel, style="Tool.TButton"
        )
        self.import_button.pack(side="left")
        self.template_button = ttk.Button(
            tools, text="下载模板", command=self.save_excel_template, style="Tool.TButton"
        )
        self.template_button.pack(side="left", padx=(8, 0))
        self.export_button = ttk.Button(
            tools, text="导出本节记录", command=self.export_history, style="Tool.TButton"
        )
        self.export_button.pack(side="left", padx=(8, 0))
        self.restore_button = ttk.Button(
            tools, text="恢复上次名单", command=self.restore_backup, style="Tool.TButton"
        )
        self.restore_button.pack(side="left", padx=(8, 0))
        ttk.Checkbutton(tools, text="音效", variable=self.sound_enabled).pack(
            side="right", pady=5
        )

        ttk.Label(
            self.root, textvariable=self.status_text, style="Status.TLabel", padding=(12, 6)
        ).pack(side="bottom", fill="x")

    def bind_shortcuts(self):
        self.root.bind("<F11>", lambda _event: self.toggle_fullscreen())
        self.root.bind("<Escape>", lambda _event: self.exit_fullscreen())
        self.root.bind("<space>", self.handle_space)

    def handle_space(self, _event):
        if self.root.focus_get() == self.class_combo:
            return None
        self.start_roll_call()
        return "break"

    def position_stage_items(self, event):
        center_x = event.width / 2
        self.stage.coords(self.stage_title, center_x, event.height * 0.18)
        self.stage.coords(self.stage_result, center_x, event.height * 0.48)
        self.stage.coords(self.stage_detail, center_x, event.height * 0.73)

    def toggle_fullscreen(self):
        self.is_fullscreen = not self.is_fullscreen
        self.root.attributes("-fullscreen", self.is_fullscreen)
        self.fullscreen_button.configure(text="退出全屏 Esc" if self.is_fullscreen else "全屏 F11")

    def exit_fullscreen(self):
        if self.is_fullscreen:
            self.is_fullscreen = False
            self.root.attributes("-fullscreen", False)
            self.fullscreen_button.configure(text="全屏 F11")

    def set_stage(self, title, result, detail, color="#0E7A6D"):
        self.stage.itemconfigure(self.stage_title, text=title)
        self.stage.itemconfigure(self.stage_result, text=result, fill=color)
        self.stage.itemconfigure(self.stage_detail, text=detail)

    def play_sound(self, sound_type, step=0):
        if not self.sound_enabled.get() or winsound is None:
            return
        try:
            if sound_type == "countdown":
                winsound.Beep(650 + step * 120, 80)
            elif sound_type == "tick":
                winsound.Beep(850 + (step % 6) * 60, 18)
            elif sound_type == "finish":
                winsound.MessageBeep(winsound.MB_ICONASTERISK)
        except RuntimeError:
            pass

    def refresh_classes(self, preferred_class=None):
        classes = self.database.get_classes()
        class_names = [class_name for class_name, _count in classes]
        self.class_combo["values"] = class_names
        if preferred_class in class_names:
            self.selected_class.set(preferred_class)
        elif self.selected_class.get() in class_names:
            pass
        elif class_names:
            self.selected_class.set(class_names[0])
        else:
            self.selected_class.set("")
        self.refresh_stats()

    def refresh_stats(self):
        class_name = self.selected_class.get()
        if not class_name:
            self.count_text.set("请先导入 Excel 名单")
            return
        students = self.database.get_students(class_name)
        state = self.database.get_roll_state(class_name)
        if state is None:
            remaining_count = len(students)
            round_no = 1
        else:
            remaining, round_no = state
            valid_numbers = {student_no for student_no, _name in students}
            remaining_count = len([number for number in remaining if number in valid_numbers])
        self.count_text.set(
            f"参与 {len(students)} 人  ·  第 {round_no} 轮  ·  剩余 {remaining_count} 人"
        )

    def on_class_changed(self, _event=None):
        if self.auto_hide_job:
            self.root.after_cancel(self.auto_hide_job)
            self.auto_hide_job = None
        self.set_stage("准备点名", "按空格键开始", "一轮内每名学生只会出现一次")
        self.roll_button.configure(text="开始点名")
        self.refresh_stats()

    def set_controls_enabled(self, enabled):
        state = "normal" if enabled else "disabled"
        self.roll_button.configure(state=state)
        self.import_button.configure(state=state)
        self.template_button.configure(state=state)
        self.export_button.configure(state=state)
        self.restore_button.configure(state=state)
        self.class_combo.configure(state="readonly" if enabled else "disabled")

    def prepare_draw(self, students, class_name):
        student_map = {student_no: name for student_no, name in students}
        state = self.database.get_roll_state(class_name)
        if state is None:
            remaining = []
            round_no = 1
        else:
            remaining, round_no = state
            remaining = [number for number in remaining if number in student_map]

        if not remaining:
            if state is not None:
                round_no += 1
            remaining = list(student_map)
            random.shuffle(remaining)
            last_selected = self.database.get_last_selected_no(class_name)
            if len(remaining) > 1 and remaining[-1] == last_selected:
                remaining[-1], remaining[-2] = remaining[-2], remaining[-1]

        selected_no = remaining.pop()
        return (selected_no, student_map[selected_no]), remaining, round_no

    def start_roll_call(self):
        if self.is_rolling:
            return
        class_name = self.selected_class.get()
        if not class_name:
            messagebox.showinfo("请先导入名单", "请先通过 Excel 导入学生名单。")
            return
        students = self.database.get_students(class_name)
        if not students:
            messagebox.showinfo("无法点名", "当前班级没有参与点名的学生。")
            return

        if self.auto_hide_job:
            self.root.after_cancel(self.auto_hide_job)
            self.auto_hide_job = None

        self.pending_student, self.pending_remaining, self.pending_round_no = (
            self.prepare_draw(students, class_name)
        )
        self.rolling_students = students
        self.is_rolling = True
        self.set_controls_enabled(False)
        self.status_text.set("点名进行中……")
        self.countdown(3)

    def countdown(self, number):
        if number > 0:
            self.set_stage("准备", str(number), "保持悬念……", "#F59E0B")
            self.play_sound("countdown", number)
            self.root.after(620, lambda: self.countdown(number - 1))
        else:
            self.roll_step = 0
            self.total_roll_steps = random.randint(34, 42)
            self.run_roll_animation()

    def run_roll_animation(self):
        if self.roll_step >= self.total_roll_steps:
            self.finish_roll_call()
            return
        student = (
            self.pending_student
            if self.roll_step == self.total_roll_steps - 1
            else random.choice(self.rolling_students)
        )
        colors = ["#0E7A6D", "#2E75B6", "#D97706", "#7C3AED", "#C2415D"]
        color = colors[self.roll_step % len(colors)]
        self.set_stage("正在抽取", masked_name(student[1]), "姓名已隐藏", color)
        if self.roll_step % 3 == 0:
            self.play_sound("tick", self.roll_step)
        self.roll_step += 1
        progress = self.roll_step / self.total_roll_steps
        delay = 42 + int(310 * (progress ** 3))
        self.root.after(delay, self.run_roll_animation)

    def finish_roll_call(self):
        class_name = self.selected_class.get()
        student_no, name = self.pending_student
        self.database.save_roll_result(
            self.session_id,
            class_name,
            self.pending_round_no,
            self.pending_remaining,
            student_no,
            name,
        )
        self.is_rolling = False
        self.set_controls_enabled(True)
        self.roll_button.configure(text="再抽一次")
        detail = f"学号：{student_no}"
        if not self.pending_remaining:
            detail += "  ·  本轮已全部点完"
        self.set_stage("本次点到", name, detail, "#0E7A6D")
        self.status_text.set(f"完整姓名将在 {AUTO_HIDE_SECONDS} 秒后自动隐藏")
        self.play_sound("finish")
        self.refresh_stats()
        self.start_confetti()
        self.flash_result(0)
        self.auto_hide_job = self.root.after(
            AUTO_HIDE_SECONDS * 1000, self.hide_result
        )

    def flash_result(self, count):
        if count >= 6:
            self.stage.itemconfigure(self.stage_result, fill="#0E7A6D")
            return
        color = "#F59E0B" if count % 2 == 0 else "#0E7A6D"
        self.stage.itemconfigure(self.stage_result, fill=color)
        self.root.after(170, lambda: self.flash_result(count + 1))

    def start_confetti(self):
        for item in self.confetti_items:
            self.stage.delete(item)
        self.confetti_items = []
        width = max(self.stage.winfo_width(), 500)
        colors = ["#FFD166", "#06D6A0", "#118AB2", "#EF476F", "#7C3AED"]
        for _index in range(55):
            x = random.randint(10, width - 10)
            y = random.randint(-180, -5)
            size = random.randint(4, 9)
            item = self.stage.create_rectangle(
                x, y, x + size, y + size, fill=random.choice(colors), outline=""
            )
            self.confetti_items.append(item)
        self.animate_confetti(0)

    def animate_confetti(self, frame):
        if frame >= 75:
            for item in self.confetti_items:
                self.stage.delete(item)
            self.confetti_items = []
            return
        for index, item in enumerate(self.confetti_items):
            self.stage.move(item, -1 if index % 2 else 1, 5 + index % 4)
        self.root.after(24, lambda: self.animate_confetti(frame + 1))

    def hide_result(self):
        self.auto_hide_job = None
        self.set_stage("结果已自动隐藏", "***", "按空格键开始下一次点名", "#607086")
        self.status_text.set("学生名单只能通过 Excel 修改")

    def import_excel(self):
        file_path = filedialog.askopenfilename(
            title="选择学生名单 Excel 文件",
            filetypes=[("Excel 工作簿", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if not file_path:
            return
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            found = find_excel_columns(worksheet)
            if found is None:
                workbook.close()
                messagebox.showerror(
                    "无法识别表头",
                    "必须包含“学号”和“姓名”两列。建议使用程序提供的模板。",
                )
                return
            header_row, columns = found
            max_column = max(columns.values())
            students = []
            seen_numbers = set()
            invalid_count = 0
            duplicate_count = 0
            inactive_count = 0

            for row in worksheet.iter_rows(min_row=header_row + 1, max_col=max_column):
                student_no = excel_cell_text(row[columns["student_no"] - 1].value)
                name = excel_cell_text(row[columns["name"] - 1].value)
                class_name = (
                    excel_cell_text(row[columns["class_name"] - 1].value)
                    if "class_name" in columns else "默认班级"
                ) or "默认班级"
                active = (
                    participation_enabled(row[columns["active"] - 1].value)
                    if "active" in columns else True
                )
                if not student_no and not name:
                    continue
                if not student_no or not name:
                    invalid_count += 1
                    continue
                if student_no in seen_numbers:
                    duplicate_count += 1
                    continue
                seen_numbers.add(student_no)
                if not active:
                    inactive_count += 1
                students.append((student_no, name, class_name, 1 if active else 0))
            workbook.close()

            if not students:
                messagebox.showwarning("没有有效数据", "Excel 中没有可导入的有效学生。")
                return
            if not messagebox.askyesno(
                "确认更新名单",
                f"将使用 Excel 中的 {len(students)} 名学生整体替换当前名单。\n"
                "旧名单会自动备份。是否继续？",
            ):
                return

            preferred_class = self.selected_class.get()
            self.database.replace_students(students)
            self.refresh_classes(preferred_class)
            self.on_class_changed()
            active_count = len(students) - inactive_count
            self.status_text.set("Excel 名单更新成功，旧名单已自动备份")
            messagebox.showinfo(
                "导入完成",
                f"导入学生：{len(students)} 名\n"
                f"参与点名：{active_count} 名\n"
                f"暂不参与：{inactive_count} 名\n"
                f"重复学号：{duplicate_count} 条\n"
                f"无效数据：{invalid_count} 条",
            )
        except Exception as error:
            messagebox.showerror("导入失败", f"无法读取该 Excel 文件：\n{error}")

    def save_excel_template(self):
        template_path = resource_path(TEMPLATE_NAME)
        destination = filedialog.asksaveasfilename(
            title="保存学生导入模板", initialfile=TEMPLATE_NAME,
            defaultextension=".xlsx", filetypes=[("Excel 工作簿", "*.xlsx")]
        )
        if not destination:
            return
        try:
            Path(destination).write_bytes(template_path.read_bytes())
            messagebox.showinfo("保存成功", "Excel 导入模板已保存。")
        except OSError as error:
            messagebox.showerror("保存失败", str(error))

    def restore_backup(self):
        backup = self.database.latest_backup()
        if backup is None:
            messagebox.showinfo("没有备份", "尚未找到可恢复的名单备份。")
            return
        if not messagebox.askyesno(
            "恢复上次名单", "确定恢复最近一次自动备份吗？当前名单也会保留为备份。"
        ):
            return
        if self.database.restore_latest_backup():
            self.refresh_classes()
            self.on_class_changed()
            self.status_text.set("已恢复上次名单")
            messagebox.showinfo("恢复成功", "已恢复最近一次名单备份。")

    def export_history(self):
        history = self.database.get_session_history(self.session_id)
        if not history:
            messagebox.showinfo("暂无记录", "本次启动后还没有完成任何点名。")
            return
        destination = filedialog.asksaveasfilename(
            title="导出本节课点名记录",
            initialfile=f"点名记录_{datetime.now():%Y%m%d_%H%M}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not destination:
            return
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "点名记录"
            sheet.append(["时间", "班级", "轮次", "学号", "姓名"])
            for selected_at, class_name, round_no, student_no, name in history:
                sheet.append([selected_at, class_name, round_no, student_no, name])
            header_fill = PatternFill("solid", fgColor="173B68")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions["A"].width = 21
            sheet.column_dimensions["B"].width = 18
            sheet.column_dimensions["C"].width = 10
            sheet.column_dimensions["D"].width = 18
            sheet.column_dimensions["E"].width = 16
            sheet.freeze_panes = "A2"
            workbook.save(destination)
            messagebox.showinfo("导出成功", "本节课点名记录已导出。")
        except Exception as error:
            messagebox.showerror("导出失败", str(error))


def main():
    root = tk.Tk()
    StudentRollCallApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
